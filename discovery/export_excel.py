"""
Export all gathered discovery/raw/<circle>/*.json runs into one shareable Excel file.

Usage:
    python export_excel.py
    python export_excel.py --out exports/my_export.xlsx
"""

import argparse
import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from discover import CIRCLES

SCRIPT_DIR = Path(__file__).parent
RAW_DIR = SCRIPT_DIR / "raw"
EXPORTS_DIR = SCRIPT_DIR / "exports"

COLUMNS = [
    "Circle", "Country", "Name", "Brief Description", "Sectors", "Geography",
    "Email", "Email Status", "Location", "Introduced By",
    "Core Fit", "Capital / Operational Orientation", "Sector Alignment Evidence",
    "Governance / Institutional Mindset", "Strategic Adjacency to TBP", "Engagement Readiness",
    "Source URLs", "Source Titles", "Discovery Date", "Search Query",
]


def load_rows() -> list[dict]:
    rows = []
    for json_path in sorted(RAW_DIR.glob("*/*.json")):
        data = json.loads(json_path.read_text(encoding="utf-8"))
        meta = data.get("run_metadata", {})
        circle_key = json_path.parent.name
        circle_label = meta.get("circle") or CIRCLES.get(circle_key, {}).get("label", circle_key)
        for p in data.get("prospects", []):
            signals = p.get("scoring_signals", {})
            sources = [
                s if isinstance(s, dict) else {"url": s, "title": None}
                for s in p.get("source_urls", [])
            ]
            rows.append({
                "Circle": circle_label,
                "Country": (meta.get("country") or "").strip().title(),
                "Name": p.get("name"),
                "Brief Description": p.get("brief_description"),
                "Sectors": "; ".join(p.get("sectors", [])),
                "Geography": p.get("geography"),
                "Email": p.get("email"),
                "Email Status": p.get("email_status"),
                "Location": p.get("location"),
                "Introduced By": p.get("introduced_by"),
                "Core Fit": signals.get("core_fit"),
                "Capital / Operational Orientation": signals.get("capital_or_operational_orientation"),
                "Sector Alignment Evidence": signals.get("sector_alignment"),
                "Governance / Institutional Mindset": signals.get("governance_institutional_mindset"),
                "Strategic Adjacency to TBP": signals.get("strategic_adjacency_tbp"),
                "Engagement Readiness": signals.get("engagement_readiness"),
                "Source URLs": "; ".join(s.get("url", "") for s in sources),
                "Source Titles": "; ".join(s.get("title") or "" for s in sources),
                "Discovery Date": meta.get("date"),
                "Search Query": meta.get("query"),
            })
    return rows


def dedupe(df: pd.DataFrame) -> pd.DataFrame:
    key = (
        df["Circle"].str.strip().str.lower() + "|"
        + df["Country"].str.strip().str.lower() + "|"
        + df["Name"].str.strip().str.lower()
    )
    before = len(df)
    df = df.loc[~key.duplicated(keep="first")].reset_index(drop=True)
    removed = before - len(df)
    if removed:
        print(f"Removed {removed} duplicate row(s) (same name within the same circle/country, kept the first occurrence)")
    return df


def write_excel(rows: list[dict], out_path: Path) -> int:
    df = dedupe(pd.DataFrame(rows, columns=COLUMNS))
    summary = df.groupby(["Circle", "Country"]).size().reset_index(name="Prospects Found")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        df.to_excel(writer, sheet_name="Prospects", index=False)

        for sheet_name, frame in [("Summary", summary), ("Prospects", df)]:
            ws = writer.sheets[sheet_name]
            for col_idx, col_name in enumerate(frame.columns, start=1):
                ws.cell(row=1, column=col_idx).font = Font(bold=True)
                max_len = max([len(str(col_name))] + [len(str(v)) for v in frame[col_name].astype(str)])
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

    return len(df)


def main():
    parser = argparse.ArgumentParser(description="Export discovery/raw/ into a shareable Excel file")
    parser.add_argument("--out", default=None, help="output path (default: exports/TBP_Prospects_<date>.xlsx)")
    args = parser.parse_args()

    rows = load_rows()
    if not rows:
        print("No prospects found under raw/<circle>/*.json — nothing to export.")
        return

    out_path = Path(args.out) if args.out else EXPORTS_DIR / f"TBP_Prospects_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
    final_count = write_excel(rows, out_path)
    combos = len(set((r["Circle"], r["Country"]) for r in rows))
    print(f"Exported {final_count} prospects across {combos} circle/country combinations to {out_path}")


if __name__ == "__main__":
    main()
